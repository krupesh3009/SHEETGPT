import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from os.path import isfile, exists, splitext
from shutil import copyfile

def load_input_workbook(file_path: str):
    wb = None

    if not exists(file_path) or not isfile(file_path):
        raise Exception("File does not exist or invalid\n")

    try:
        wb = load_workbook(file_path, read_only=True)
    except Exception as e:
        print(e)
        raise Exception("Failed to load workbook. Is the file format correct (xlsx/xlsm/xltx/xltm)?\n")

    return wb

def create_output_book(file_path: str, result_path: str, overwrite=False) -> str:
    result_book = None

    try:
        copyfile(file_path, result_path)
        result_book = load_workbook(result_path)
    except Exception as e:
        print(e)
        raise Exception("\nFailed to create output worksheet.\n")

    return result_book

def generate_result_path(file_path: str) -> str:
    pre, ext = splitext(file_path)
    return pre + "_output" + ext

def open_file_dialog():
    file_path = filedialog.askopenfilename()
    if file_path:
        try:
            input_wb = load_input_workbook(file_path)
            result_path = generate_result_path(file_path)
            create_output_book(file_path, result_path)
            # Do something with the output workbook, like displaying it or further processing
        except Exception as e:
            print("Error:", e)

# Create a Tkinter window
window = tk.Tk()

# Add a button to open file dialog
open_button = tk.Button(window, text="Open Workbook", command=open_file_dialog)
open_button.pack()

# Run the Tkinter event loop
window.mainloop()
